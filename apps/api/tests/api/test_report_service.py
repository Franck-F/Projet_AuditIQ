"""TDD: report_service.get_or_build_excel — cache-or-build."""
import io
import uuid

import httpx as _httpx
import pytest
import respx as _respx
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import async_sessionmaker

from app.core.db import Base, make_engine
from app.integrations.storage import MemoryStorage
from app.models import Organization, Report, User
from app.schemas.audit import AuditCreate
from app.services import audit_service, dataset_service, report_service


def _recruitment_csv() -> bytes:
    rows = ["genre,decision"]
    rows += ["Hommes,oui"] * 100 + ["Hommes,non"] * 100
    rows += ["Femmes,oui"] * 72 + ["Femmes,non"] * 128
    return ("\n".join(rows) + "\n").encode()


@pytest.fixture
async def ctx(tmp_path):
    """Yield (session_maker, org_id, audit_id) for a completed M1 audit."""
    eng = make_engine(f"sqlite+aiosqlite:///{tmp_path / 'a.db'}")
    async with eng.begin() as c:
        await c.run_sync(Base.metadata.create_all)
    sm = async_sessionmaker(eng, expire_on_commit=False)

    store = MemoryStorage()
    async with sm() as s:
        org = Organization(name="acme")
        s.add(org)
        await s.flush()
        user = User(id=uuid.uuid4(), org_id=org.id, email="a@acme.fr")
        s.add(user)
        await s.commit()
        org_id, uid = org.id, user.id

    async with sm() as s:
        ds = await dataset_service.create_dataset(
            s, store, org_id=org_id, user_id=uid,
            filename="r.csv", raw=_recruitment_csv(), retention_days=30,
        )
        out = await audit_service.run_m1_audit(
            s, store, org_id=org_id, user_id=uid,
            body=AuditCreate(
                dataset_id=ds.id, title="Recrutement",
                protected_attribute="genre", decision_column="decision",
                favorable_value="oui",
            ),
            llm_provider=None,
        )
        audit_id = out.id

    yield sm, org_id, audit_id
    await eng.dispose()


async def test_get_or_build_excel_builds_then_caches(ctx):
    sm, org_id, audit_id = ctx
    store = MemoryStorage()
    async with sm() as session:
        b1, name = await report_service.get_or_build_excel(
            session, store, audit_id, org_id=org_id
        )
        assert name.endswith(".xlsx")
        assert load_workbook(io.BytesIO(b1)).sheetnames
        rows = (
            await session.execute(select(Report).where(Report.audit_id == audit_id))
        ).scalars().all()
        assert len(rows) == 1 and rows[0].format == "xlsx"
    async with sm() as session:
        b2, _ = await report_service.get_or_build_excel(
            session, store, audit_id, org_id=org_id
        )
        assert b2 == b1
        rows = (
            await session.execute(select(Report).where(Report.audit_id == audit_id))
        ).scalars().all()
        assert len(rows) == 1


async def test_get_or_build_pdf_builds_then_caches(ctx, monkeypatch):
    sm, org_id, audit_id = ctx
    monkeypatch.setenv("PDF_SERVICE_URL", "http://pdf:8080")
    monkeypatch.setenv("PDF_SERVICE_SECRET", "shh")
    from app.core.config import get_settings

    get_settings.cache_clear()
    store = MemoryStorage()
    try:
        with _respx.mock:
            r = _respx.post("http://pdf:8080/render").mock(
                return_value=_httpx.Response(200, content=b"%PDF-1.7 x")
            )
            async with sm() as session:
                b1, name = await report_service.get_or_build_pdf(
                    session, store, audit_id, org_id=org_id
                )
                assert name.endswith(".pdf")
                assert b1 == b"%PDF-1.7 x"
                rows = (
                    await session.execute(
                        select(Report).where(Report.audit_id == audit_id)
                    )
                ).scalars().all()
                assert any(x.format == "pdf" for x in rows)
            async with sm() as session:
                b2, _n = await report_service.get_or_build_pdf(
                    session, store, audit_id, org_id=org_id
                )
            assert b2 == b1
            assert r.call_count == 1
    finally:
        get_settings.cache_clear()
