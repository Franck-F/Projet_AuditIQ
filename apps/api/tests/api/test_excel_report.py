import datetime
import io
import uuid

from openpyxl import load_workbook

from app.reporting.excel import build_excel_report
from app.schemas.audit import (
    AuditOut,
    ClusterStatOut,
    GroupStatOut,
    InterpretationOut,
    IntersectionalCellOut,
    IntersectionalOut,
    M1MetricsOut,
    M2MetricsOut,
    RecommendationOut,
)

_NOW = datetime.datetime(2026, 5, 16, tzinfo=datetime.timezone.utc)


def _interp() -> InterpretationOut:
    return InterpretationOut(
        narrative="Texte.",
        ai_act_anchors=["AI Act, article 10"],
        disclaimers=["Aide à l'analyse."],
        provider="fallback",
        model="deterministic",
    )


def _m1_audit() -> AuditOut:
    return AuditOut(
        id=uuid.uuid4(), code="AUD-2026-001", title="Recrutement", status="done",
        module="M1", dataset_id=uuid.uuid4(), protected_attribute="genre",
        decision_column="embauche", favorable_value="oui", privileged_value=None,
        created_at=_NOW, completed_at=_NOW,
        metrics=M1MetricsOut(
            groups=[GroupStatOut(value="F", n=100, favorable=30,
                                 selection_rate=0.3, disparate_impact=0.6)],
            reference_value="H", disparate_impact=0.6,
            demographic_parity_diff=0.2, worst_group="F", verdict="fail",
            risk_score=80, warnings=[],
        ),
        interpretation=_interp(), pre_check=["Déséquilibre groupe F."],
        config=None,
    )


def _m2_audit() -> AuditOut:
    return AuditOut(
        id=uuid.uuid4(), code="AUD-2026-002", title="Détection", status="done",
        module="M2", dataset_id=uuid.uuid4(), protected_attribute=None,
        decision_column="embauche", favorable_value="oui", privileged_value=None,
        created_at=_NOW, completed_at=_NOW,
        metrics=M2MetricsOut(
            n=240, k=2, global_positive_rate=0.5, chi2=88.1, p_value=0.0001,
            dof=1,
            clusters=[ClusterStatOut(id=0, n=120, positive_rate=0.1,
                                     deviation_pp=-40, is_deviant=True,
                                     top_features=[])],
            deviant_cluster_ids=[0], verdict="fail", risk_score=88, warnings=[],
        ),
        interpretation=_interp(), pre_check=[], config={"k": 2},
    )


def _text(b: bytes) -> str:
    wb = load_workbook(io.BytesIO(b))
    return " ".join(
        str(c.value)
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for c in row
        if c.value is not None
    )


def test_build_excel_m1_has_sheets_and_not_a_certificate():
    b = build_excel_report(_m1_audit())
    assert isinstance(b, bytes) and len(b) > 0
    wb = load_workbook(io.BytesIO(b))
    assert "Résumé" in wb.sheetnames
    assert "Détail" in wb.sheetnames
    assert "Conformité" in wb.sheetnames
    t = _text(b)
    assert "n'est pas un certificat" in t
    assert "AUD-2026-001" in t
    assert "Disparate Impact" in t


def test_build_excel_m2_renders_cluster_detail():
    t = _text(build_excel_report(_m2_audit()))
    assert "AUD-2026-002" in t
    assert "p-value" in t or "Khi-deux" in t
    assert "n'est pas un certificat" in t


def _m3_audit() -> AuditOut:
    from app.schemas.audit import (
        CategoryStatOut,
        DivergentExampleOut,
        M3MetricsOut,
    )
    return AuditOut(
        id=uuid.uuid4(), code="AUD-2026-030", title="Chatbot", status="done",
        module="M3", dataset_id=None, protected_attribute=None,
        decision_column=None, favorable_value=None, privileged_value=None,
        created_at=_NOW, completed_at=_NOW,
        metrics=M3MetricsOut(
            categories=[CategoryStatOut(name="genre", length_gap=0.4,
                        sentiment_gap=0.2, refusal_rate=0.5, score=0.55,
                        verdict="warn")],
            global_score=0.55, verdict="warn", risk_score=55,
            divergent_examples=[DivergentExampleOut(category="genre",
                prompt_id="g1", variant_a="m", variant_b="f",
                excerpt_a="long", excerpt_b="Je ne peux pas",
                reason="refus")],
            n_pairs=1, n_calls_failed=0, warnings=[]),
        interpretation=_interp(), pre_check=[], config={"lang": "fr"},
    )


def test_build_excel_m3_section():
    import io

    from openpyxl import load_workbook

    from app.reporting.excel import build_excel_report

    wb = load_workbook(io.BytesIO(build_excel_report(_m3_audit())))
    text = " ".join(
        str(c.value) for ws in wb.worksheets for row in ws.iter_rows()
        for c in row if c.value is not None
    )
    assert "AUD-2026-030" in text
    assert "genre" in text
    assert "n'est pas un certificat" in text


def _m1_audit_with_eo() -> AuditOut:
    return AuditOut(
        id=uuid.uuid4(), code="AUD-2026-010", title="Recrutement EO", status="done",
        module="M1", dataset_id=uuid.uuid4(), protected_attribute="genre",
        decision_column="embauche", favorable_value="oui", privileged_value=None,
        created_at=_NOW, completed_at=_NOW,
        metrics=M1MetricsOut(
            groups=[
                GroupStatOut(value="F", n=100, favorable=30,
                             selection_rate=0.3, disparate_impact=0.6,
                             tpr=0.5, fpr=0.4),
                GroupStatOut(value="H", n=100, favorable=50,
                             selection_rate=0.5, disparate_impact=1.0,
                             tpr=0.9, fpr=0.1),
            ],
            reference_value="H", disparate_impact=0.6,
            demographic_parity_diff=0.2, worst_group="F", verdict="fail",
            risk_score=80, warnings=[],
            equal_opportunity_diff=0.4,
            equalized_odds_diff=0.4,
            demographic_parity_verdict="warn",
            equal_opportunity_verdict="fail",
            equalized_odds_verdict="fail",
            truelabel_reason=None,
        ),
        interpretation=_interp(), pre_check=["Déséquilibre groupe F."],
        config=None,
    )


def test_excel_m1_eo_section_present_and_absent():
    b_with = build_excel_report(_m1_audit_with_eo())
    b_without = build_excel_report(_m1_audit())

    t_with = _text(b_with)
    t_without = _text(b_without)

    # EO section present when fields are set
    assert "Equal Opportunity" in t_with
    assert "Equalized Odds" in t_with
    # per-group TPR values present
    assert "0.5" in t_with
    assert "0.9" in t_with

    # Existing M1 cells present in BOTH
    assert "Disparate Impact" in t_with
    assert "Disparate Impact" in t_without

    # No EO section when fields are absent
    assert "Equal Opportunity" not in t_without
    assert "Equalized Odds" not in t_without


def test_excel_m1_with_truelabel_reason():
    audit = _m1_audit_with_eo()
    # rebuild with a truelabel_reason set
    assert audit.metrics is not None
    m = audit.metrics
    assert isinstance(m, M1MetricsOut)
    audit2 = AuditOut(
        id=audit.id, code="AUD-2026-011", title=audit.title, status=audit.status,
        module=audit.module, dataset_id=audit.dataset_id,
        protected_attribute=audit.protected_attribute,
        decision_column=audit.decision_column, favorable_value=audit.favorable_value,
        privileged_value=audit.privileged_value,
        created_at=audit.created_at, completed_at=audit.completed_at,
        metrics=M1MetricsOut(
            groups=m.groups, reference_value=m.reference_value,
            disparate_impact=m.disparate_impact,
            demographic_parity_diff=m.demographic_parity_diff,
            worst_group=m.worst_group, verdict=m.verdict,
            risk_score=m.risk_score, warnings=m.warnings,
            equal_opportunity_diff=None,
            equalized_odds_diff=None,
            demographic_parity_verdict=None,
            equal_opportunity_verdict=None,
            equalized_odds_verdict=None,
            truelabel_reason="Non calculable : moins de 2 groupes.",
        ),
        interpretation=audit.interpretation,
        pre_check=audit.pre_check, config=audit.config,
    )
    t = _text(build_excel_report(audit2))
    assert "Non calculable" in t
    assert "Disparate Impact" in t


def _intersectional_out() -> IntersectionalOut:
    return IntersectionalOut(
        cells=[
            IntersectionalCellOut(
                primary_value="femme", secondary_value="etr",
                n=20, favorable=4, selection_rate=0.2,
                disparate_impact=0.4, verdict="fail",
            ),
            IntersectionalCellOut(
                primary_value="femme", secondary_value="fr",
                n=80, favorable=32, selection_rate=0.4,
                disparate_impact=0.8, verdict="warn",
            ),
            IntersectionalCellOut(
                primary_value="homme", secondary_value="etr",
                n=30, favorable=15, selection_rate=0.5,
                disparate_impact=1.0, verdict="pass",
            ),
            IntersectionalCellOut(
                primary_value="homme", secondary_value="fr",
                n=120, favorable=60, selection_rate=0.5,
                disparate_impact=1.0, verdict="pass",
            ),
        ],
        reference_primary="homme",
        reference_secondary="fr",
        worst_primary="femme",
        worst_secondary="etr",
        disparate_impact=0.4,
        demographic_parity_diff=0.3,
        verdict="fail",
        risk_score=75,
        marginal_di=[0.82, 0.9],
    )


def _m1_audit_with_pairwise() -> AuditOut:
    from app.schemas.audit import MarginalOut
    ix = _intersectional_out()
    # set attribute names for the pairwise entry
    ix_with_attrs = ix.model_copy(
        update={"primary_attribute": "genre", "secondary_attribute": "origine"}
    )
    return AuditOut(
        id=uuid.uuid4(), code="AUD-2026-050", title="Recrutement Pairwise", status="done",
        module="M1", dataset_id=uuid.uuid4(), protected_attribute="genre",
        decision_column="embauche", favorable_value="oui", privileged_value=None,
        created_at=_NOW, completed_at=_NOW,
        metrics=M1MetricsOut(
            groups=[GroupStatOut(value="F", n=100, favorable=30,
                                 selection_rate=0.3, disparate_impact=0.6)],
            reference_value="H", disparate_impact=0.6,
            demographic_parity_diff=0.2, worst_group="F", verdict="fail",
            risk_score=80, warnings=[],
            marginals=[
                MarginalOut(
                    attribute="genre",
                    groups=[GroupStatOut(value="F", n=100, favorable=30,
                                         selection_rate=0.3, disparate_impact=0.6)],
                    reference_value="H", disparate_impact=0.6,
                    demographic_parity_diff=0.2, worst_group="F",
                    verdict="fail", risk_score=80,
                ),
                MarginalOut(
                    attribute="origine",
                    groups=[GroupStatOut(value="etr", n=50, favorable=10,
                                         selection_rate=0.2, disparate_impact=0.4)],
                    reference_value="fr", disparate_impact=0.82,
                    demographic_parity_diff=0.18, worst_group="etr",
                    verdict="warn", risk_score=50,
                ),
            ],
            pairwise=[ix_with_attrs],
        ),
        interpretation=_interp(), pre_check=["Déséquilibre groupe F."],
        config=None,
    )


def test_excel_m1_pairwise_section_present_and_absent():
    b_with = build_excel_report(_m1_audit_with_pairwise())
    b_without = build_excel_report(_m1_audit())

    t_with = _text(b_with)
    t_without = _text(b_without)

    # Pairwise section present: crossed subgroup values
    assert "etr" in t_with
    assert "femme" in t_with
    assert "homme" in t_with
    # pair heading (attribut names) and worst cell labels
    assert "Pire sous-groupe" in t_with or "femme" in t_with
    # marginal DI values from pairwise entry
    assert "0.82" in t_with

    # Per-attribute marginal sections
    assert "genre" in t_with
    assert "origine" in t_with
    assert "Attribut protégé" in t_with

    # Existing M1 cells present in BOTH
    assert "Disparate Impact" in t_with
    assert "Disparate Impact" in t_without

    # No pairwise section when absent (no pairwise list)
    assert "intersectionnel" not in t_without.lower()
    assert "etr" not in t_without


def _interp_with_recommendations(
    recs: list[RecommendationOut],
) -> InterpretationOut:
    return InterpretationOut(
        narrative="Texte.",
        ai_act_anchors=["AI Act, article 10"],
        disclaimers=["Aide à l'analyse."],
        provider="fallback",
        model="deterministic",
        recommendations=recs,
    )


def _sample_audit_with_recommendations(recs: list[RecommendationOut]) -> AuditOut:
    return AuditOut(
        id=uuid.uuid4(), code="AUD-2026-REC", title="Recrutement Recos", status="done",
        module="M1", dataset_id=uuid.uuid4(), protected_attribute="genre",
        decision_column="embauche", favorable_value="oui", privileged_value=None,
        created_at=_NOW, completed_at=_NOW,
        metrics=M1MetricsOut(
            groups=[GroupStatOut(value="F", n=100, favorable=30,
                                 selection_rate=0.3, disparate_impact=0.6)],
            reference_value="H", disparate_impact=0.6,
            demographic_parity_diff=0.2, worst_group="F", verdict="fail",
            risk_score=80, warnings=[],
        ),
        interpretation=_interp_with_recommendations(recs),
        pre_check=[],
        config=None,
    )


def test_excel_includes_recommendations_sheet_when_present() -> None:
    """Excel workbook gains a 'Recommandations' sheet when interpretation has recos."""
    audit = _sample_audit_with_recommendations([
        RecommendationOut(title="Re-collecter données", detail="Détail 1.", priority="high"),
        RecommendationOut(title="Audit features", detail="Détail 2.", priority="medium"),
    ])
    data = build_excel_report(audit)
    wb = load_workbook(io.BytesIO(data))
    assert "Recommandations" in wb.sheetnames
    ws = wb["Recommandations"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("#", "Priorité", "Action", "Détail")
    assert rows[1] == (1, "Action prioritaire", "Re-collecter données", "Détail 1.")
    assert rows[2] == (2, "À planifier", "Audit features", "Détail 2.")


def test_excel_omits_recommendations_sheet_when_empty() -> None:
    """No recos → no sheet."""
    audit = _sample_audit_with_recommendations([])
    data = build_excel_report(audit)
    wb = load_workbook(io.BytesIO(data))
    assert "Recommandations" not in wb.sheetnames
