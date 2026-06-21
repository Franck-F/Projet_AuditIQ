"""Pure Excel audit report: build_excel_report(AuditOut) -> bytes."""
from __future__ import annotations

import io
from collections.abc import Sequence

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from app.reporting.format import (
    module_label,
    p_value_display,
    status_label,
    verdict_label,
)
from app.schemas.audit import (
    RECOMMENDATION_CATEGORY_LABELS,
    AuditOut,
    M1MetricsOut,
    M2MetricsOut,
    M3MetricsOut,
    RecommendationOut,
)

_HORIZON_LABEL = {
    "immediat": "Immédiat", "court_terme": "Court terme", "continu": "Continu",
}

_VERDICT_BADGE_FR = {
    "fail": "🔴 Risque élevé",
    "warn": "🟠 Vigilance",
    "pass": "🟢 Risque faible",
}
_NOT_A_CERTIFICATE = (
    "Ce rapport est une aide à l'analyse documentée : il n'est pas un "
    "certificat de conformité ni un avis juridique."
)
_AI_ACT_MAP = [
    ("Article 9 — système de gestion des risques", "Risque agrégé / verdict"),
    ("Article 10 — qualité et gouvernance des données", "Pré-vérifications / groupes"),
    ("Article 11 — documentation technique", "Ce rapport (trace documentée)"),
]
_FR_LAW = (
    "Références droit français : Code du travail L.1132-1 ; Défenseur des "
    "droits ; CNIL ; ACPR (selon le contexte d'usage)."
)


def _rows(ws: Worksheet, rows: Sequence[Sequence[object]]) -> None:
    for r in rows:
        ws.append(r)


def _write_recommendations_sheet(
    wb: Workbook, recs: list[RecommendationOut]
) -> None:
    """Recommandations déployeur : catégorie, priorité graduée, responsable,
    horizon, référence légale et sous-étapes."""
    if not recs:
        return
    ws = wb.create_sheet("Recommandations")
    ws.append([
        "#", "Priorité", "Catégorie", "Action", "Pourquoi (constat)",
        "Responsable", "Horizon", "Référence légale", "Sous-étapes",
    ])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for idx, rec in enumerate(recs, start=1):
        ws.append([
            idx,
            rec.priority_level,
            RECOMMENDATION_CATEGORY_LABELS.get(rec.category, rec.category),
            rec.title,
            rec.rationale or rec.detail,
            rec.owner,
            _HORIZON_LABEL.get(rec.horizon, rec.horizon),
            rec.legal_ref or "—",
            " ; ".join(rec.steps) if rec.steps else "—",
        ])
    ws.append([])
    ws.append([
        "Actions à la portée du déployeur de l'outil — par priorité "
        "(1 = haute)."
    ])


def build_excel_report(audit: AuditOut) -> bytes:
    wb = Workbook()
    summary = wb.active
    summary.title = "Résumé"
    m = audit.metrics
    verdict = m.verdict if m is not None else "—"
    risk = m.risk_score if m is not None else "—"
    _rows(
        summary,
        [
            ["Rapport d'audit de fairness — AuditIQ"],
            [],
            ["Audit", audit.code or str(audit.id)],
            ["Titre", audit.title],
            ["Module", module_label(audit.module)],
            ["Statut", status_label(audit.status)],
            ["Verdict", _VERDICT_BADGE_FR.get(str(verdict), str(verdict))],
            ["Score de risque", f"{risk}/100"],
            [],
            [_NOT_A_CERTIFICATE],
        ],
    )

    detail = wb.create_sheet("Détail")
    if m is None:
        _rows(detail, [["Résultats indisponibles pour cet audit."]])
    elif isinstance(m, M2MetricsOut):
        _rows(
            detail,
            [
                [module_label("M2")],
                ["Test du Khi-deux (p-value)", p_value_display(m.p_value)],
                ["χ²", m.chi2, "ddl", m.dof],
                ["Taux favorable global", m.global_positive_rate],
                ["Clusters déviants", f"{len(m.deviant_cluster_ids)} / {m.k}"],
                [],
                ["Cluster", "Effectif", "Taux fav.", "Écart (pts)", "Déviant"],
            ],
        )
        for c in m.clusters:
            detail.append(
                [c.id, c.n, c.positive_rate, c.deviation_pp,
                 "oui" if c.is_deviant else "non"]
            )
    elif isinstance(m, M3MetricsOut):
        _rows(
            detail,
            [
                [module_label("M3")],
                ["Score global", m.global_score],
                ["Verdict", verdict_label(m.verdict)],
                ["Paires", m.n_pairs, "Appels échoués", m.n_calls_failed],
                [],
                ["Catégorie", "Écart long.", "Écart sent.", "Taux refus",
                 "Score", "Verdict"],
            ],
        )
        for cat in m.categories:
            detail.append([cat.name, cat.length_gap, cat.sentiment_gap,
                           cat.refusal_rate, cat.score,
                           verdict_label(cat.verdict)])
    elif isinstance(m, M1MetricsOut):
        # Mono-attribut : le résumé module et la section « Attribut protégé »
        # afficheraient deux fois les mêmes métriques — on ne garde que la
        # section par attribut.
        single_attr = len(m.marginals) == 1
        if single_attr:
            _rows(detail, [[module_label("M1")]])
        else:
            _rows(
                detail,
                [
                    [module_label("M1")],
                    ["Disparate Impact", m.disparate_impact],
                    ["Demographic Parity (écart)", m.demographic_parity_diff],
                    ["Groupe le plus défavorisé", m.worst_group],
                    ["Référence", m.reference_value],
                    [],
                    ["Groupe", "Effectif", "Favorables", "Taux", "DI"],
                ],
            )
            for g in m.groups:
                detail.append(
                    [g.value, g.n, g.favorable, g.selection_rate,
                     g.disparate_impact]
                )
        if m.equal_opportunity_diff is not None or m.truelabel_reason is not None:
            _rows(detail, [[]])
            if m.truelabel_reason is not None:
                _rows(detail, [["Note vérité-terrain", m.truelabel_reason]])
            if m.equal_opportunity_diff is not None:
                _rows(
                    detail,
                    [
                        ["Métriques vérité-terrain (Equal Opportunity / Equalized Odds)"],
                        ["Equal Opportunity (écart TPR)",
                         m.equal_opportunity_diff,
                         "Verdict EO",
                         verdict_label(m.equal_opportunity_verdict)
                         if m.equal_opportunity_verdict else "—"],
                        ["Equalized Odds (écart max TPR/FPR)",
                         m.equalized_odds_diff,
                         "Verdict Eq. Odds",
                         verdict_label(m.equalized_odds_verdict)
                         if m.equalized_odds_verdict else "—"],
                        ["Demographic Parity (verdict)",
                         verdict_label(m.demographic_parity_verdict)
                         if m.demographic_parity_verdict else "—"],
                        [],
                        ["Groupe", "TPR", "FPR"],
                    ],
                )
                for g in m.groups:
                    if g.tpr is not None or g.fpr is not None:
                        detail.append(
                            [g.value,
                             g.tpr if g.tpr is not None else "—",
                             g.fpr if g.fpr is not None else "—"]
                        )
                _rows(
                    detail,
                    [
                        [],
                        ["Note normative",
                         "DP, Equal Opportunity et Equalized Odds ne peuvent "
                         "être satisfaits simultanément — tout choix de "
                         "métrique est un choix normatif, pas seulement "
                         "technique."],
                    ],
                )
        for marg in m.marginals:
            _rows(
                detail,
                [
                    [],
                    [f"Attribut protégé : {marg.attribute}"],
                    ["Disparate Impact", marg.disparate_impact],
                    ["Demographic Parity (écart)", marg.demographic_parity_diff],
                    ["Ratio de parité (indicateur complémentaire)",
                     marg.demographic_parity_ratio],
                    ["Groupe le plus défavorisé", marg.worst_group],
                    ["Référence", marg.reference_value],
                    ["Verdict", verdict_label(marg.verdict)],
                    ["Score de risque", marg.risk_score],
                ],
            )
            if marg.equal_opportunity_ratio is not None:
                _rows(
                    detail,
                    [
                        ["Ratio Equal Opportunity (indicateur complémentaire)",
                         marg.equal_opportunity_ratio],
                        ["Ratio Equalized Odds (indicateur complémentaire)",
                         marg.equalized_odds_ratio if marg.equalized_odds_ratio is not None else "—"],
                    ],
                )
            # Per-group table — add FNR/Accuracy/Precision columns when GT present
            has_gt_rates = any(
                g.fnr is not None for g in marg.groups
            )
            if has_gt_rates:
                _rows(detail, [[], ["Groupe", "Effectif", "Favorables",
                                     "Taux", "DI", "FNR", "Accuracy",
                                     "Precision"]])
                for g in marg.groups:
                    detail.append(
                        [g.value, g.n, g.favorable, g.selection_rate,
                         g.disparate_impact,
                         g.fnr if g.fnr is not None else "—",
                         g.accuracy if g.accuracy is not None else "—",
                         g.precision if g.precision is not None else "—"]
                    )
            else:
                _rows(detail, [[], ["Groupe", "Effectif", "Favorables",
                                     "Taux", "DI"]])
                for g in marg.groups:
                    detail.append(
                        [g.value, g.n, g.favorable, g.selection_rate,
                         g.disparate_impact]
                    )
            # Mono-attribut : le bloc vérité-terrain global couvre déjà ces
            # lignes — éviter le doublon.
            if marg.equal_opportunity_diff is not None and not (
                single_attr and m.equal_opportunity_diff is not None
            ):
                _rows(
                    detail,
                    [
                        [],
                        ["Equal Opportunity (écart TPR)",
                         marg.equal_opportunity_diff,
                         "Verdict EO",
                         verdict_label(marg.equal_opportunity_verdict)
                         if marg.equal_opportunity_verdict else "—"],
                        ["Equalized Odds (écart max TPR/FPR)",
                         marg.equalized_odds_diff,
                         "Verdict Eq. Odds",
                         verdict_label(marg.equalized_odds_verdict)
                         if marg.equalized_odds_verdict else "—"],
                    ],
                )
            if marg.warnings:
                for w in marg.warnings:
                    detail.append(["Avertissement", w])
            if marg.truelabel_reason is not None and not (
                single_attr and m.truelabel_reason is not None
            ):
                _rows(detail, [["Note vérité-terrain", marg.truelabel_reason]])
        for ix in m.pairwise:
            pair_title = (
                f"{ix.primary_attribute} × {ix.secondary_attribute}"
                if ix.primary_attribute
                else "Analyse intersectionnelle"
            )
            _rows(
                detail,
                [
                    [],
                    [f"Croisement : {pair_title}"],
                    ["Disparate Impact intersectionnel", ix.disparate_impact],
                    ["Demographic Parity (écart intersectionnel)",
                     ix.demographic_parity_diff],
                    ["Ratio de parité intersectionnel "
                     "(indicateur complémentaire)",
                     ix.demographic_parity_ratio],
                    ["Verdict intersectionnel", verdict_label(ix.verdict)],
                    ["Pire sous-groupe (primaire)", ix.worst_primary],
                    ["Pire sous-groupe (secondaire)", ix.worst_secondary],
                    ["DI marginal (attribut primaire)", ix.marginal_di[0]
                     if len(ix.marginal_di) > 0 else "—"],
                    ["DI marginal (attribut secondaire)", ix.marginal_di[1]
                     if len(ix.marginal_di) > 1 else "—"],
                ],
            )
            if ix.equal_opportunity_ratio is not None:
                _rows(
                    detail,
                    [
                        ["Ratio Equal Opportunity intersectionnel "
                         "(indicateur complémentaire)",
                         ix.equal_opportunity_ratio],
                        ["Ratio Equalized Odds intersectionnel "
                         "(indicateur complémentaire)",
                         ix.equalized_odds_ratio if ix.equalized_odds_ratio is not None else "—"],
                    ],
                )
            _rows(
                detail,
                [
                    [],
                    ["Matrice croisée : sous-groupes"],
                    ["Groupe primaire", "Groupe secondaire", "Effectif",
                     "Favorables", "Taux", "DI", "Verdict"],
                ],
            )
            for cell in ix.cells:
                detail.append([
                    cell.primary_value, cell.secondary_value, cell.n,
                    cell.favorable, cell.selection_rate,
                    cell.disparate_impact, verdict_label(cell.verdict),
                ])
            if ix.equal_opportunity_diff is not None:
                _rows(
                    detail,
                    [
                        [],
                        ["Equal Opportunity intersectionnel (écart TPR)",
                         ix.equal_opportunity_diff,
                         "Verdict EO intersect.",
                         verdict_label(ix.equal_opportunity_verdict)
                         if ix.equal_opportunity_verdict else "—"],
                        ["Equalized Odds intersectionnel (écart max TPR/FPR)",
                         ix.equalized_odds_diff,
                         "Verdict Eq. Odds intersect.",
                         verdict_label(ix.equalized_odds_verdict)
                         if ix.equalized_odds_verdict else "—"],
                    ],
                )
            if ix.warnings:
                for w in ix.warnings:
                    detail.append(["Avertissement", w])
            if ix.reason is not None:
                _rows(detail, [["Note", ix.reason]])

    references = wb.create_sheet("Références")
    _rows(
        references,
        [
            ["Mise en regard AI Act"],
            ["Article", "Élément du rapport"],
            *[[a, b] for a, b in _AI_ACT_MAP],
            [],
            [_FR_LAW],
            [],
            ["Pré-vérifications"],
            *([[w] for w in audit.pre_check] or [["Aucune."]]),
            [],
            ["Limites"],
            *(
                [[d] for d in audit.interpretation.disclaimers]
                if audit.interpretation is not None
                else [["—"]]
            ),
            [],
            [_NOT_A_CERTIFICATE],
        ],
    )

    _write_recommendations_sheet(
        wb,
        audit.interpretation.recommendations if audit.interpretation is not None else [],
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
