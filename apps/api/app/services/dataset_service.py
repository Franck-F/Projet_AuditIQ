from __future__ import annotations

import csv
import datetime
import io
import json
import uuid

import openpyxl
import pandas as pd
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.audit_engine import run_dataset_analysis
from app.core.errors import APIError, NotFoundError
from app.integrations.storage import Storage
from app.models import Dataset
from app.schemas.dataset import DatasetAnalysisOut

# Extensions / types MIME reconnus comme classeurs Excel (.xlsx uniquement).
# Le format binaire historique .xls n'est PAS pris en charge (openpyxl ne le
# lit pas) : on demande à l'utilisateur de l'enregistrer au format .xlsx.
_XLSX_EXTENSIONS = (".xlsx",)
_XLSX_MIME_TYPES = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)


def _is_xlsx(filename: str, content_type: str | None) -> bool:
    """Détecte un classeur Excel .xlsx par extension OU type MIME."""
    name = (filename or "").lower()
    if name.endswith(_XLSX_EXTENSIONS):
        return True
    if content_type is None:
        return False
    return content_type.lower() in _XLSX_MIME_TYPES


def _xlsx_cell_to_text(value: object) -> str:
    """Convertit une valeur de cellule en texte pour le CSV normalisé."""
    if value is None:
        return ""
    if isinstance(value, bool):
        # Évite "True"/"False" : on garde une représentation stable.
        return "1" if value else "0"
    if isinstance(value, datetime.datetime):
        # Sans composante horaire utile -> date seule, sinon ISO.
        if value.hour == 0 and value.minute == 0 and value.second == 0:
            return value.date().isoformat()
        return value.isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        # 1.0 -> "1" (les entiers Excel remontent en float).
        return str(int(value))
    return str(value)


def _xlsx_to_csv(raw: bytes) -> bytes:
    """Convertit la PREMIÈRE feuille d'un .xlsx en CSV UTF-8 normalisé.

    Lecture en mode `read_only` + `data_only` : on récupère les valeurs
    calculées des cellules sans jamais exécuter ni interpréter de formule
    ni de macro. La 1re ligne non vide sert d'en-têtes ; le séparateur est
    la virgule et l'échappement (virgules, guillemets, sauts de ligne)
    est délégué au module `csv` standard.
    """
    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(raw), read_only=True, data_only=True
        )
    except Exception as exc:  # openpyxl/zipfile lèvent divers types  # noqa: BLE001
        raise APIError(
            "Le fichier Excel est illisible ou endommagé.",
            title="Invalid Dataset",
            status=422,
        ) from exc

    try:
        if not workbook.sheetnames:
            raise APIError(
                "Le fichier Excel semble vide (aucune feuille).",
                title="Invalid Dataset",
                status=422,
            )
        # On ne traite que la première feuille du classeur.
        sheet = workbook[workbook.sheetnames[0]]
        rows_iter = sheet.iter_rows(values_only=True)

        # Recherche de la première ligne non entièrement vide = en-têtes.
        header: list[str] | None = None
        for row in rows_iter:
            if row is None:
                continue
            if any(cell is not None and str(cell).strip() != "" for cell in row):
                header = [_xlsx_cell_to_text(cell).strip() for cell in row]
                break

        if header is None:
            raise APIError(
                "Le fichier Excel semble vide (aucune donnée).",
                title="Invalid Dataset",
                status=422,
            )
        # Retire les colonnes de queue sans en-tête (cellules vides à droite).
        while header and header[-1] == "":
            header.pop()
        if not header or all(c == "" for c in header):
            raise APIError(
                "Le fichier Excel ne contient pas d'en-têtes de colonnes.",
                title="Invalid Dataset",
                status=422,
            )
        width = len(header)

        buffer = io.StringIO(newline="")
        writer = csv.writer(buffer, lineterminator="\n")
        writer.writerow(header)

        has_data = False
        for row in rows_iter:
            if row is None:
                continue
            cells = [_xlsx_cell_to_text(cell) for cell in row]
            # Aligne sur la largeur des en-têtes (tronque / complète).
            cells = (cells + [""] * width)[:width]
            if all(cell.strip() == "" for cell in cells):
                continue  # ignore les lignes entièrement vides
            writer.writerow(cells)
            has_data = True

        if not has_data:
            raise APIError(
                "Le fichier Excel ne contient aucune ligne de données.",
                title="Invalid Dataset",
                status=422,
            )
    finally:
        workbook.close()

    return buffer.getvalue().encode("utf-8")


def _parse_csv(raw: bytes) -> tuple[list[str], int]:
    try:
        df = pd.read_csv(io.BytesIO(raw))
    except Exception as exc:  # pandas raises many types for bad input  # noqa: BLE001
        raise APIError(
            "Fichier CSV illisible.",
            title="Invalid Dataset",
            status=422,
        ) from exc
    if df.shape[1] == 0 or df.empty:
        raise APIError(
            "Le CSV ne contient aucune donnée exploitable.",
            title="Invalid Dataset",
            status=422,
        )
    return [str(c) for c in df.columns], int(len(df))


async def create_dataset(
    session: AsyncSession,
    storage: Storage,
    *,
    org_id: uuid.UUID,
    user_id: uuid.UUID,
    filename: str,
    raw: bytes,
    retention_days: int,
    content_type: str | None = None,
) -> Dataset:
    # Les classeurs Excel (.xlsx) sont convertis en CSV normalisé dès
    # l'ingestion : tout le pipeline aval (analyse de colonnes, moteur,
    # stockage) reste inchangé car il ne manipule que du CSV.
    if _is_xlsx(filename, content_type):
        raw = _xlsx_to_csv(raw)
    columns, row_count = _parse_csv(raw)
    dataset_id = uuid.uuid4()
    storage_path = f"{org_id}/{dataset_id}.csv"
    await storage.upload(storage_path, raw, "text/csv")
    now = datetime.datetime.now(tz=datetime.timezone.utc)
    dataset = Dataset(
        id=dataset_id,
        org_id=org_id,
        uploaded_by=user_id,
        filename=filename,
        storage_path=storage_path,
        row_count=row_count,
        columns=columns,
        status="ready",
        expires_at=now + datetime.timedelta(days=retention_days),
    )
    session.add(dataset)
    await session.commit()
    return dataset


async def get_dataset(
    session: AsyncSession, dataset_id: uuid.UUID, *, org_id: uuid.UUID
) -> Dataset:
    dataset = (
        await session.execute(
            select(Dataset).where(
                Dataset.id == dataset_id, Dataset.org_id == org_id
            )
        )
    ).scalar_one_or_none()
    if dataset is None:
        raise NotFoundError("Jeu de données introuvable.")
    return dataset


async def get_or_build_analysis(
    session: AsyncSession,
    storage: Storage,
    dataset_id: uuid.UUID,
    *,
    org_id: uuid.UUID,
) -> DatasetAnalysisOut:
    """Return cached analysis or compute, cache, and return it.

    RLS enforced via org_id scope. Raises NotFoundError for cross-org access.
    """
    dataset = await get_dataset(session, dataset_id, org_id=org_id)
    if dataset.analysis_cache:
        return DatasetAnalysisOut.model_validate(dataset.analysis_cache)
    raw = await storage.download(dataset.storage_path)
    df = pd.read_csv(io.BytesIO(raw))
    analysis = run_dataset_analysis(df)
    out = DatasetAnalysisOut.from_engine(analysis)
    dataset.analysis_cache = json.loads(out.model_dump_json())
    await session.commit()
    return out
